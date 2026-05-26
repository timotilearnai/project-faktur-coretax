import os
from openpyxl import load_workbook

# File daftar invoice
wb_daftar = load_workbook("daftar invoice.xlsx")
ws_daftar = wb_daftar.active

# Ambil data dari kolom I dan K
data_invoice = []
for row in ws_daftar.iter_rows(min_row=2, max_col=11):  # Kolom K = ke-11
    no_faktur = row[8].value  # Kolom I = index 8
    kode_template = row[10].value  # Kolom K = index 10

    if no_faktur and kode_template:
        data_invoice.append((str(no_faktur).strip(), str(kode_template).strip()))

# Siapkan folder output
os.makedirs("output_faktur", exist_ok=True)

# Proses semua faktur
for faktur, kode_temp in data_invoice:
    nama_template = f"template{kode_temp}.xlsx"

    if not os.path.exists(nama_template):
        print(f"❌ Template {nama_template} tidak ditemukan! Lewati {faktur}")
        continue

    try:
        wb_template = load_workbook(nama_template)
        if "Faktur" not in wb_template.sheetnames:
            print(f"⚠️ Sheet 'Faktur' tidak ditemukan di {nama_template}, dilewati.")
            continue

        ws_faktur = wb_template["Faktur"]
        ws_faktur["D1"] = faktur

        output_path = os.path.join("output_faktur", f"{faktur}.xlsx")
        wb_template.save(output_path)

        print(f"✅ {faktur} dibuat dari {nama_template}")

    except Exception as e:
        print(f"❌ Error saat proses {faktur}: {e}")

print("\n🎉 Sukses Asisten Ai telah selesai timoti !")
